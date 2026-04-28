from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
import re
import uuid

import psycopg
from openpyxl import load_workbook


ENV_PATH = Path(r"C:\Users\paul.loja\PYPROYECTOS\dashboard_v2\.env.local")
WORKBOOK_PATH = Path(r"C:\Users\paul.loja\PYPROYECTOS\dashboard_v2\docs\bodega\propuesta-categorizacion-productos-base.xlsx")

UNIT_REF_TABLE = "public.bodega_ref_unit_id_core_scd2"
UNIT_DIM_TABLE = "public.bodega_dim_unit_profile_scd2"
CATEGORY_REF_TABLE = "public.bodega_ref_category_id_core_scd2"
CATEGORY_DIM_TABLE = "public.bodega_dim_category_profile_scd2"
PRODUCT_REF_TABLE = "public.bodega_ref_product_id_core_scd2"
PRODUCT_DIM_TABLE = "public.bodega_dim_product_profile_scd2"
PRODUCT_USAGE_TABLE = "public.bodega_bridge_product_usage_scd2"


@dataclass
class PreparedProduct:
    source_sheet: str
    product_code: str
    product_name: str
    unit_code: str
    active_component_mode: str
    active_component_name: str | None
    tipo: str
    familia: str
    subfamilia: str
    assignments: list[str]


def load_env(file_path: Path) -> dict[str, str]:
    result: dict[str, str] = {}
    for line in file_path.read_text(encoding="utf-8").splitlines():
        trimmed = line.strip()
        if not trimmed or trimmed.startswith("#") or "=" not in trimmed:
            continue
        key, value = trimmed.split("=", 1)
        result[key.strip()] = value.strip()
    return result


def clean(value: object) -> str:
    if value is None:
      return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def normalize_code(value: str) -> str:
    return clean(value).upper()


def normalize_name(value: str) -> str:
    return clean(value)


def entity_id(prefix: str, key: str) -> str:
    return f"{prefix}_{uuid.uuid5(uuid.NAMESPACE_URL, key)}"


def record_id() -> str:
    return str(uuid.uuid4())


def map_unit_name(unit_code: str) -> tuple[str, str, int]:
    normalized = normalize_code(unit_code)
    mapping = {
        "UN": ("Unidad", "Unidad", 0),
        "UND": ("Unidad", "Unidad", 0),
        "PR": ("Par", "Unidad", 0),
        "CJ": ("Caja", "Unidad", 0),
        "PK": ("Paquete", "Unidad", 0),
        "SA": ("Saco", "Unidad", 0),
        "RO": ("Rollo", "Unidad", 0),
        "RM": ("Rollo madre", "Unidad", 0),
        "KG": ("Kilogramo", "Peso", 2),
        "GR": ("Gramo", "Peso", 2),
        "LT": ("Litro", "Volumen", 2),
        "CC": ("Centimetro cubico", "Volumen", 2),
        "GL": ("Galon", "Volumen", 2),
        "GA": ("Galon", "Volumen", 2),
        "M3": ("Metro cubico", "Volumen", 3),
        "P3": ("Pie cubico", "Volumen", 3),
        "PIE3": ("Pie cubico", "Volumen", 3),
        "MT": ("Metro", "Longitud", 2),
    }
    if normalized in mapping:
        name, dimension, precision = mapping[normalized]
        return name, dimension, precision
    return normalized or "Unidad", "Unidad", 0


def normalize_unit_code(raw_unit: str) -> str:
    normalized = normalize_code(raw_unit)
    aliases = {
        "UND": "UN",
        "UNID": "UN",
    }
    return aliases.get(normalized, normalized)


def is_valid_unit_code(unit_code: str) -> bool:
    if not unit_code:
        return False
    if len(unit_code) > 12:
        return False
    return True


def source_priority(source_sheet: str) -> int:
    return 0 if source_sheet == "Presupuesto" else 1


def load_products_from_workbook(path: Path) -> tuple[list[PreparedProduct], dict[str, int]]:
    wb = load_workbook(path, read_only=True)
    ws = wb["Productos"]
    rows = ws.iter_rows(min_row=1, values_only=True)
    headers = [clean(value) for value in next(rows)]
    idx = {header: position for position, header in enumerate(headers)}

    grouped: dict[str, list[dict[str, object]]] = defaultdict(list)
    summary = {
        "rows_seen": 0,
        "rows_skipped_invalid_code": 0,
        "rows_skipped_invalid_unit": 0,
        "rows_grouped_duplicates": 0,
    }

    for row in rows:
        summary["rows_seen"] += 1
        product_code = normalize_code(row[idx["codigo"]])
        if not product_code or product_code == "0":
            summary["rows_skipped_invalid_code"] += 1
            continue

        raw_unit = clean(row[idx["unidad"]])
        unit_code = normalize_unit_code(raw_unit)
        if not is_valid_unit_code(unit_code):
            summary["rows_skipped_invalid_unit"] += 1
            continue

        grouped[product_code].append(
            {
                "source_sheet": clean(row[idx["source_sheet"]]),
                "product_code": product_code,
                "product_name": normalize_name(row[idx["descripcion"]]),
                "unit_code": unit_code,
                "active_component_mode": normalize_code(row[idx["active_component_mode_propuesto"]]).lower() or "na",
                "active_component_name": normalize_name(row[idx["ingrd_activo"]]) or None,
                "tipo": normalize_name(row[idx["tipo_propuesto"]]),
                "familia": normalize_name(row[idx["familia_propuesta"]]),
                "subfamilia": normalize_name(row[idx["subfamilia_propuesta"]]),
                "assignments": [
                    normalize_code(row[idx["activity_id_1"]]),
                    normalize_code(row[idx["activity_id_2"]]),
                    normalize_code(row[idx["activity_id_3"]]),
                ],
            }
        )

    products: list[PreparedProduct] = []
    for product_code, options in grouped.items():
        if len(options) > 1:
            summary["rows_grouped_duplicates"] += len(options) - 1
        selected = sorted(
            options,
            key=lambda item: (
                source_priority(str(item["source_sheet"])),
                len(str(item["product_name"])),
                len(str(item["unit_code"])),
            ),
            reverse=False,
        )[0]
        products.append(
            PreparedProduct(
                source_sheet=str(selected["source_sheet"]),
                product_code=product_code,
                product_name=str(selected["product_name"]),
                unit_code=str(selected["unit_code"]),
                active_component_mode=str(selected["active_component_mode"]) if str(selected["active_component_mode"]) in {"applies", "na"} else "na",
                active_component_name=selected["active_component_name"],
                tipo=str(selected["tipo"]),
                familia=str(selected["familia"]),
                subfamilia=str(selected["subfamilia"]),
                assignments=[activity_id for activity_id in selected["assignments"] if activity_id],
            )
        )

    return products, summary


def fetch_valid_activity_ids(cur) -> set[str]:
    cur.execute(
        """
        select distinct trim(activity_id::text) as activity_id
        from slv.prod_dim_activity_profile_scd2
        where is_current = true
          and is_valid = true
          and nullif(trim(activity_id::text), '') is not null
        """
    )
    return {normalize_code(row[0]) for row in cur.fetchall()}


def main() -> None:
    env = load_env(ENV_PATH)
    products, workbook_summary = load_products_from_workbook(WORKBOOK_PATH)

    conn = psycopg.connect(
        host=env["DATABASE_HOST"],
        port=int(env["DATABASE_PORT"]),
        dbname=env.get("CAMP_DATABASE_NAME", "db_camp"),
        user=env["DATABASE_USER"],
        password=env["DATABASE_PASSWORD"],
        sslmode="require" if env.get("DATABASE_SSL") == "true" else "prefer",
    )

    source_conn = psycopg.connect(
        host=env["DATABASE_HOST"],
        port=int(env["DATABASE_PORT"]),
        dbname=env.get("DATABASE_NAME", "datalakehouse"),
        user=env["DATABASE_USER"],
        password=env["DATABASE_PASSWORD"],
        sslmode="require" if env.get("DATABASE_SSL") == "true" else "prefer",
    )

    now = "2026-04-27T12:00:00"
    run_id = "bodega_seed_workbook_v1"
    actor_id = "corex_bodega_seed"
    change_reason = "SEED_FROM_BODEGA_WORKBOOK"

    with conn, conn.cursor() as cur, source_conn, source_conn.cursor() as source_cur:
        for table in [UNIT_DIM_TABLE, CATEGORY_DIM_TABLE, PRODUCT_DIM_TABLE, PRODUCT_USAGE_TABLE]:
            cur.execute(f"select count(*) from {table} where is_current = true")
            total = int(cur.fetchone()[0] or 0)
            if total > 0:
                raise RuntimeError(f"El seed inicial requiere tablas vacias. {table} ya tiene {total} registros vigentes.")

        valid_activity_ids = fetch_valid_activity_ids(source_cur)

        invalid_assignment_count = 0
        for product in products:
            product.assignments = [activity_id for activity_id in product.assignments if activity_id in valid_activity_ids]
            invalid_assignment_count += 3 - len(product.assignments) if len(product.assignments) < 3 else 0

        products_by_name: dict[str, list[PreparedProduct]] = defaultdict(list)
        for product in products:
            products_by_name[normalize_name(product.product_name).lower()].append(product)
        for same_name_products in products_by_name.values():
            if len(same_name_products) > 1:
                for product in same_name_products:
                    product.product_name = f"{product.product_name} ({product.product_code})"

        used_units = sorted({product.unit_code for product in products})
        unit_id_by_code: dict[str, str] = {}
        for unit_code in used_units:
            unit_id = entity_id("bunit", f"bodega-unit:{unit_code}")
            unit_id_by_code[unit_code] = unit_id
            unit_name, unit_dimension, unit_precision = map_unit_name(unit_code)

            cur.execute(
                f"""
                insert into {UNIT_REF_TABLE} (
                  record_id, unit_id, valid_from, valid_to, is_current, is_valid, loaded_at, run_id, actor_id, change_reason
                ) values (%s, %s, %s, null, true, true, %s, %s, %s, %s)
                """,
                (record_id(), unit_id, now, now, run_id, actor_id, change_reason),
            )
            cur.execute(
                f"""
                insert into {UNIT_DIM_TABLE} (
                  record_id, unit_id, valid_from, valid_to, is_current, unit_code, unit_name, unit_symbol,
                  unit_dimension, decimal_precision, is_active, is_valid, loaded_at, run_id, actor_id, change_reason
                ) values (%s, %s, %s, null, true, %s, %s, %s, %s, %s, true, true, %s, %s, %s, %s)
                """,
                (record_id(), unit_id, now, unit_code, unit_name, unit_code, unit_dimension, unit_precision, now, run_id, actor_id, change_reason),
            )

        type_names = sorted({product.tipo for product in products})
        category_id_by_path: dict[tuple[str, str, str], str] = {}
        type_id_by_name: dict[str, str] = {}
        family_id_by_name: dict[tuple[str, str], str] = {}

        for type_index, tipo in enumerate(type_names, start=1):
            category_id = entity_id("bcat", f"bodega-category:type:{tipo}")
            type_id_by_name[tipo] = category_id
            cur.execute(
                f"""
                insert into {CATEGORY_REF_TABLE} (
                  record_id, category_id, valid_from, valid_to, is_current, is_valid, loaded_at, run_id, actor_id, change_reason
                ) values (%s, %s, %s, null, true, true, %s, %s, %s, %s)
                """,
                (record_id(), category_id, now, now, run_id, actor_id, change_reason),
            )
            cur.execute(
                f"""
                insert into {CATEGORY_DIM_TABLE} (
                  record_id, category_id, valid_from, valid_to, is_current, category_code, category_name, category_level,
                  parent_category_id, sort_order, category_description, is_active, is_valid, loaded_at, run_id, actor_id, change_reason
                ) values (%s, %s, %s, null, true, %s, %s, 'type', null, %s, null, true, true, %s, %s, %s, %s)
                """,
                (record_id(), category_id, now, normalize_code(tipo)[:50], tipo, type_index, now, run_id, actor_id, change_reason),
            )

        families_by_type: dict[str, list[str]] = defaultdict(list)
        subfamilies_by_family: dict[tuple[str, str], list[str]] = defaultdict(list)
        for product in products:
            if product.familia not in families_by_type[product.tipo]:
                families_by_type[product.tipo].append(product.familia)
            if product.subfamilia not in subfamilies_by_family[(product.tipo, product.familia)]:
                subfamilies_by_family[(product.tipo, product.familia)].append(product.subfamilia)

        for tipo, families in sorted(families_by_type.items()):
            for family_index, familia in enumerate(sorted(families), start=1):
                category_id = entity_id("bcat", f"bodega-category:family:{tipo}:{familia}")
                family_id_by_name[(tipo, familia)] = category_id
                parent_category_id = type_id_by_name[tipo]
                cur.execute(
                    f"""
                    insert into {CATEGORY_REF_TABLE} (
                      record_id, category_id, valid_from, valid_to, is_current, is_valid, loaded_at, run_id, actor_id, change_reason
                    ) values (%s, %s, %s, null, true, true, %s, %s, %s, %s)
                    """,
                    (record_id(), category_id, now, now, run_id, actor_id, change_reason),
                )
                cur.execute(
                    f"""
                    insert into {CATEGORY_DIM_TABLE} (
                      record_id, category_id, valid_from, valid_to, is_current, category_code, category_name, category_level,
                      parent_category_id, sort_order, category_description, is_active, is_valid, loaded_at, run_id, actor_id, change_reason
                    ) values (%s, %s, %s, null, true, %s, %s, 'family', %s, %s, null, true, true, %s, %s, %s, %s)
                    """,
                    (record_id(), category_id, now, normalize_code(familia)[:50], familia, parent_category_id, family_index, now, run_id, actor_id, change_reason),
                )

        for (tipo, familia), subfamilies in sorted(subfamilies_by_family.items()):
            for subfamily_index, subfamilia in enumerate(sorted(subfamilies), start=1):
                category_id = entity_id("bcat", f"bodega-category:subfamily:{tipo}:{familia}:{subfamilia}")
                category_id_by_path[(tipo, familia, subfamilia)] = category_id
                parent_category_id = family_id_by_name[(tipo, familia)]
                cur.execute(
                    f"""
                    insert into {CATEGORY_REF_TABLE} (
                      record_id, category_id, valid_from, valid_to, is_current, is_valid, loaded_at, run_id, actor_id, change_reason
                    ) values (%s, %s, %s, null, true, true, %s, %s, %s, %s)
                    """,
                    (record_id(), category_id, now, now, run_id, actor_id, change_reason),
                )
                cur.execute(
                    f"""
                    insert into {CATEGORY_DIM_TABLE} (
                      record_id, category_id, valid_from, valid_to, is_current, category_code, category_name, category_level,
                      parent_category_id, sort_order, category_description, is_active, is_valid, loaded_at, run_id, actor_id, change_reason
                    ) values (%s, %s, %s, null, true, %s, %s, 'subfamily', %s, %s, null, true, true, %s, %s, %s, %s)
                    """,
                    (record_id(), category_id, now, normalize_code(subfamilia)[:50], subfamilia, parent_category_id, subfamily_index, now, run_id, actor_id, change_reason),
                )

        for product in products:
            product_id = entity_id("bprod", f"bodega-product:{product.product_code}")
            category_id = category_id_by_path[(product.tipo, product.familia, product.subfamilia)]
            unit_id = unit_id_by_code[product.unit_code]

            cur.execute(
                f"""
                insert into {PRODUCT_REF_TABLE} (
                  record_id, product_id, valid_from, valid_to, is_current, is_valid, loaded_at, run_id, actor_id, change_reason
                ) values (%s, %s, %s, null, true, true, %s, %s, %s, %s)
                """,
                (record_id(), product_id, now, now, run_id, actor_id, change_reason),
            )
            cur.execute(
                f"""
                insert into {PRODUCT_DIM_TABLE} (
                  record_id, product_id, valid_from, valid_to, is_current, product_code, product_name,
                  product_description, base_unit_id, category_id, active_component_mode, active_component_name, is_active,
                  is_valid, loaded_at, run_id, actor_id, change_reason
                ) values (%s, %s, %s, null, true, %s, %s, %s, %s, %s, %s, %s, true, true, %s, %s, %s, %s)
                """,
                (
                    record_id(),
                    product_id,
                    now,
                    product.product_code,
                    product.product_name,
                    f"Fuente inicial: {product.source_sheet}",
                    unit_id,
                    category_id,
                    product.active_component_mode,
                    product.active_component_name if product.active_component_mode == "applies" else None,
                    now,
                    run_id,
                    actor_id,
                    change_reason,
                ),
            )

            for branch_order, activity_id in enumerate(product.assignments, start=1):
                cur.execute(
                    f"""
                    insert into {PRODUCT_USAGE_TABLE} (
                      record_id, product_id, valid_from, valid_to, is_current, branch_order, activity_id,
                      is_valid, loaded_at, run_id, actor_id, change_reason
                    ) values (%s, %s, %s, null, true, %s, %s, true, %s, %s, %s, %s)
                    """,
                    (record_id(), product_id, now, branch_order, activity_id, now, run_id, actor_id, change_reason),
                )

        print({
            "units_seeded": len(used_units),
            "categories_seeded": len(type_id_by_name) + len(family_id_by_name) + len(category_id_by_path),
            "products_seeded": len(products),
            "products_with_assignments": sum(1 for product in products if product.assignments),
            "products_without_assignments": sum(1 for product in products if not product.assignments),
            "rows_seen": workbook_summary["rows_seen"],
            "rows_skipped_invalid_code": workbook_summary["rows_skipped_invalid_code"],
            "rows_skipped_invalid_unit": workbook_summary["rows_skipped_invalid_unit"],
            "rows_collapsed_as_duplicates": workbook_summary["rows_grouped_duplicates"],
        })

    conn.close()
    source_conn.close()


if __name__ == "__main__":
    main()
